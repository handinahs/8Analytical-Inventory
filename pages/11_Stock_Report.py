from io import BytesIO
from datetime import date
import sqlite3

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.ui import enable_select_all_inputs, translate_columns, render_brand_header

render_brand_header("📑 Relatório de Estoque")

enable_select_all_inputs()

DB_PATH = "estoque.db"


def load_products_with_stock():
    conn = sqlite3.connect(DB_PATH)

    query = """
    SELECT
        p.codigo AS Code,
        p.descricao AS Description,
        p.unidade AS Unit,
        COALESCE(p.posicao, '') AS Location,
        p.qtd_inicial
        + COALESCE((
            SELECT SUM(quantidade)
            FROM movimentacoes
            WHERE codigo_produto = p.codigo
            AND tipo = 'IN'
            AND status = 1
        ), 0)
        - COALESCE((
            SELECT SUM(quantidade)
            FROM movimentacoes
            WHERE codigo_produto = p.codigo
            AND tipo = 'OUT'
            AND status = 1
        ), 0) AS Current_Stock
    FROM produtos p
    WHERE p.ativo = 1
    ORDER BY p.descricao
    """

    df = pd.read_sql(query, conn)
    conn.close()
    return df


def build_excel(report_data, items_df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatorio de Estoque"

    blue = "1F4E79"
    light_blue = "D9E6F2"
    pale_yellow = "FFF2CC"
    white = "FFFFFF"
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"] = "RELATÓRIO DE ESTOQUE"
    ws["A1"].fill = PatternFill("solid", fgColor=blue)
    ws["A1"].font = Font(color=white, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:D3")
    ws["A3"] = f"Responsável:{report_data['responsible']}"
    ws["E3"] = "Data:"
    ws["F3"] = report_data["report_date"].strftime("%d-%b-%y")

    ws.merge_cells("A4:D4")
    ws["A4"] = f"Nº Relatório: {report_data['report_number']}"
    ws.merge_cells("E4:F4")
    ws["E4"] = f"Destino:{report_data['destination']}"

    ws.merge_cells("A5:F5")
    ws["A5"] = f"Observação:{report_data['note']}"

    headers = ["ITEM", "CÓDIGO", "DESCRIÇÃO", "UN", "POSIÇÃO", "EM ESTOQUE"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor=light_blue)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for index, row in items_df.iterrows():
        excel_row = index + 7
        values = [
            index + 1,
            row["Code"],
            row["Description"],
            row["Unit"],
            row["Location"] or "-",
            row["Current_Stock"],
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left")

        ws.cell(row=excel_row, column=6).fill = PatternFill("solid", fgColor=pale_yellow)

    for row in range(7 + len(items_df), 17):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if col == 6:
                cell.fill = PatternFill("solid", fgColor=pale_yellow)

    widths = [7, 18, 58, 8, 18, 18]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row in range(1, 17):
        ws.row_dimensions[row].height = 22

    for row in ws.iter_rows(min_row=3, max_row=5, min_col=1, max_col=6):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=light_blue)
            cell.border = border
            cell.font = Font(bold=True if cell.column in (1, 5) else False)

    ws["F3"].alignment = Alignment(horizontal="right")

    wb.save(output)
    output.seek(0)
    return output


products = load_products_with_stock()

if "stock_report_items" not in st.session_state:
    st.session_state.stock_report_items = []

st.subheader("Dados do Relatório")

col1, col2, col3, col4 = st.columns([2, 1, 1, 1])

with col1:
    responsible = st.text_input("Responsável", value="Emerson Durante")

with col2:
    destination = st.text_input("Destino", value="")

with col3:
    report_number = st.text_input("Nº Relatório", value="")

with col4:
    report_date = st.date_input("Data", value=date.today())

note = st.text_input("Observação", value="Consulta de saldo em estoque")

st.divider()
st.subheader("Adicionar Item")

if products.empty:
    st.warning("Nenhum produto cadastrado.")
    st.stop()

product_options = {
    f"{row['Code']} - {row['Description']}": row
    for _, row in products.iterrows()
}

with st.form("stock_report_item_form", clear_on_submit=False):
    selected_product = st.selectbox("Produto", list(product_options.keys()))

    submitted = st.form_submit_button("Adicionar Item")

    if submitted:
        product = product_options[selected_product]
        existing_codes = [item["Code"] for item in st.session_state.stock_report_items]

        if product["Code"] in existing_codes:
            st.info("Este item já está no relatório.")
        else:
            st.session_state.stock_report_items.append({
                "Code": product["Code"],
                "Description": product["Description"],
                "Unit": product["Unit"],
                "Location": product["Location"],
                "Current_Stock": product["Current_Stock"],
            })
            st.success("Item adicionado com sucesso!")
            st.rerun()

items_df = pd.DataFrame(st.session_state.stock_report_items)

st.divider()
st.subheader("Itens do Relatório")

if items_df.empty:
    st.info("Adicione produtos para montar o relatório de estoque.")
else:
    display_df = items_df.copy()
    display_df.insert(0, "Item", range(1, len(display_df) + 1))

    st.dataframe(
        translate_columns(display_df),
        use_container_width=True,
        hide_index=True,
    )

    item_options = {
        f"{index + 1} - {row['Code']} - {row['Description']}": index
        for index, row in items_df.iterrows()
    }

    selected_item = st.selectbox("Item para Remover", list(item_options.keys()))

    col1, col2, col3 = st.columns([1, 1, 1])

    with col1:
        if st.button("Remover Item", use_container_width=True):
            remove_index = item_options[selected_item]
            st.session_state.stock_report_items.pop(remove_index)
            st.rerun()

    with col2:
        if st.button("Limpar Relatório", use_container_width=True):
            st.session_state.stock_report_items = []
            st.rerun()

    with col3:
        report_data = {
            "responsible": responsible,
            "destination": destination,
            "report_number": report_number,
            "report_date": report_date,
            "note": note,
        }
        excel_file = build_excel(report_data, items_df)

        file_number = report_number.strip() or report_date.strftime("%Y%m%d")

        st.download_button(
            "Baixar Excel",
            data=excel_file,
            file_name=f"relatorio_estoque_{file_number}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )



