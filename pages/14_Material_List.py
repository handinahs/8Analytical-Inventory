from datetime import date
from io import BytesIO
import sqlite3

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils.ui import enable_select_all_inputs, render_brand_header, translate_columns

render_brand_header("📄 LM - Lista de Materiais")

enable_select_all_inputs()

DB_PATH = "estoque.db"


def load_products_with_stock():
    conn = sqlite3.connect(DB_PATH)

    query = """
    SELECT
        p.codigo AS Code,
        p.descricao AS Description,
        p.unidade AS Unit,
        COALESCE(p.posicao, '') AS Location,
        p.qtd_inicial
        + COALESCE((
            SELECT SUM(quantidade)
            FROM movimentacoes
            WHERE codigo_produto = p.codigo
            AND tipo = 'IN'
            AND status = 1
        ), 0)
        - COALESCE((
            SELECT SUM(quantidade)
            FROM movimentacoes
            WHERE codigo_produto = p.codigo
            AND tipo = 'OUT'
            AND status = 1
        ), 0) AS Stock
    FROM produtos p
    WHERE p.ativo = 1
    ORDER BY p.descricao
    """

    df = pd.read_sql(query, conn)
    conn.close()
    return df


def set_border(ws, cell_range, border):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border


def build_excel(lm_data, items_df):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "LM"

    black = "000000"
    white = "FFFFFF"
    light_peach = "FCE4D6"
    light_gray = "EDEDED"
    thin = Side(style="thin", color=black)
    medium = Side(style="medium", color=black)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium_border = Border(left=thin, right=thin, top=medium, bottom=medium)

    ws.merge_cells("A1:H2")
    ws["A1"] = "LISTA DE MATERIAIS | REQUISIÇÃO DE COMPRAS"
    ws["A1"].font = Font(name="Times New Roman", size=20, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G3:H3")
    ws["G3"] = f"LM | RC nº: {lm_data['lm_number']}"
    ws["G3"].font = Font(bold=True)
    ws["G3"].alignment = Alignment(horizontal="right")

    ws.merge_cells("G4:H4")
    ws["G4"] = f"Data de emissão: {lm_data['issue_date'].strftime('%d/%m/%Y')}"
    ws["G4"].font = Font(bold=True)
    ws["G4"].alignment = Alignment(horizontal="right")

    ws.merge_cells("G5:H5")
    ws["G5"] = f"Prazo necessário de atendimento: {lm_data['deadline']}"
    ws["G5"].font = Font(bold=True)
    ws["G5"].alignment = Alignment(horizontal="right")

    ws.merge_cells("A6:F6")
    ws["A6"] = "SOLICITAÇÃO DE MATERIAIS E CHECKLIST"
    ws.merge_cells("G6:H6")
    ws["G6"] = "REQUISIÇÃO DE COMPRAS"

    for cell in ("A6", "G6"):
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal="center")
        ws[cell].border = medium_border

    headers = [
        "ITEM",
        "CÓDIGO",
        "DESCRIÇÃO / ESPECIFICAÇÃO DO MATERIAL",
        "UNID.",
        "QTDE.",
        "LISTA DE\nVERIFICAÇÃO?",
        "QTDE FALTANTE",
        "PC Nº",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    start_row = 8
    min_item_rows = 8

    for index in range(max(len(items_df), min_item_rows)):
        excel_row = start_row + index

        if index < len(items_df):
            row = items_df.iloc[index]
            values = [
                index + 1,
                row["Code"],
                row["Description"],
                row["Unit"],
                row["Quantity"],
                row["Checklist"],
                row["Missing_Quantity"],
                row["Purchase_Order"],
            ]
        else:
            values = [index + 1, "", "", "", "", "", "", ""]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col, value=value)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left" if col == 3 else "center",
                vertical="center",
                wrap_text=True,
            )
            if col in (7, 8):
                cell.fill = PatternFill("solid", fgColor=light_gray)

    footer_row = start_row + max(len(items_df), min_item_rows)

    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=2)
    ws.cell(footer_row, 1, "CLIENTE | OBRA | PV")
    ws.merge_cells(start_row=footer_row, start_column=3, end_row=footer_row, end_column=8)
    ws.cell(footer_row, 3, lm_data["client_project"])

    for col in range(1, 9):
        cell = ws.cell(footer_row, col)
        cell.fill = PatternFill("solid", fgColor=light_peach)
        cell.font = Font(bold=True)
        cell.border = medium_border
        cell.alignment = Alignment(vertical="center")

    info_rows = [
        ("EMITIDO POR:", lm_data["issued_by"], "VISTO:", ""),
        ("APROVADO POR:", "", "DATA:", ""),
        ("RECEBIDO POR ALMOXARIFADO EM:", "", "VISTO:", ""),
        ("RECEBIDO POR COMPRAS EM:", "", "VISTO:", ""),
    ]

    for row_offset, (left_label, left_value, right_label, right_value) in enumerate(info_rows, start=1):
        row_number = footer_row + row_offset
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=4)
        ws.merge_cells(start_row=row_number, start_column=5, end_row=row_number, end_column=8)
        ws.cell(row_number, 1, f"{left_label} {left_value}".strip())
        ws.cell(row_number, 5, f"{right_label} {right_value}".strip())

        for col in range(1, 9):
            cell = ws.cell(row_number, col)
            cell.border = border
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(vertical="center")

    comments_row = footer_row + 5
    ws.merge_cells(start_row=comments_row, start_column=1, end_row=comments_row, end_column=8)
    ws.cell(comments_row, 1, f"Comentários | Observações: {lm_data['comments']}")
    ws.cell(comments_row, 1).font = Font(bold=True, size=9)
    ws.cell(comments_row, 1).alignment = Alignment(horizontal="center")
    set_border(ws, f"A{comments_row}:H{comments_row}", border)

    widths = [7, 18, 62, 10, 12, 18, 17, 15]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_number in range(1, comments_row + 1):
        ws.row_dimensions[row_number].height = 23

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[7].height = 32

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output)
    output.seek(0)
    return output


products = load_products_with_stock()

if "material_list_items" not in st.session_state:
    st.session_state.material_list_items = []

st.subheader("Dados da LM")

col1, col2, col3, col4 = st.columns([1, 1, 1.5, 1.5])

with col1:
    lm_number = st.text_input("LM | RC nº", value="0440")

with col2:
    issue_date = st.date_input("Data de emissão", value=date.today())

with col3:
    deadline = st.text_input("Prazo necessário de atendimento", value="IMEDIATO")

with col4:
    issued_by = st.text_input("Emitido por", value="Emerson Durante")

client_project = st.text_input("Cliente | Obra | PV", value="")
comments = st.text_area("Comentários | Observações", value="")

st.divider()
st.subheader("Adicionar Material")

if products.empty:
    st.warning("Nenhum produto cadastrado.")
    st.stop()

product_options = {
    f"{row['Code']} - {row['Description']}": row
    for _, row in products.iterrows()
}

selected_product = st.selectbox("Material", list(product_options.keys()))
product = product_options[selected_product]

stock_value = product["Stock"] or 0
location_value = product["Location"] if str(product["Location"]).strip() else "Sem posição"
st.info(
    f"Estoque atual: {stock_value:,.2f} {product['Unit'] or ''} | Posição: {location_value}"
)

with st.form("material_list_item_form", clear_on_submit=False):
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])

    with col1:
        quantity = st.number_input("Quantidade", min_value=1.0, step=1.0)

    with col2:
        checklist = st.checkbox("Lista de verificação?")

    with col3:
        missing_quantity = st.number_input("Qtde faltante", min_value=0.0, step=1.0)

    with col4:
        purchase_order = st.text_input("PC nº")

    submitted = st.form_submit_button("Adicionar Material")

    if submitted:
        st.session_state.material_list_items.append({
            "Code": product["Code"],
            "Description": product["Description"],
            "Unit": product["Unit"],
            "Quantity": quantity,
            "Checklist": "SIM" if checklist else "",
            "Missing_Quantity": missing_quantity,
            "Purchase_Order": purchase_order,
            "Current_Stock": stock_value,
            "Location": location_value,
        })
        st.success("Material adicionado com sucesso!")
        st.rerun()

items_df = pd.DataFrame(st.session_state.material_list_items)

st.divider()
st.subheader("Itens da LM")

if items_df.empty:
    st.info("Adicione materiais para montar a LM.")
else:
    display_df = items_df.copy()
    display_df.insert(0, "Item", range(1, len(display_df) + 1))
    display_df = display_df.rename(columns={
        "Checklist": "Lista de Verificação?",
        "Missing_Quantity": "Qtde Faltante",
        "Purchase_Order": "PC nº",
        "Current_Stock": "Estoque Atual",
    })

    st.dataframe(
        translate_columns(display_df),
        use_container_width=True,
        hide_index=True,
    )

    item_options = {
        f"{index + 1} - {row['Code']} - {row['Description']}": index
        for index, row in items_df.iterrows()
    }

    selected_item = st.selectbox("Material para Remover", list(item_options.keys()))

    col1, col2, col3 = st.columns([1, 1, 1])

    with col1:
        if st.button("Remover Material", use_container_width=True):
            remove_index = item_options[selected_item]
            st.session_state.material_list_items.pop(remove_index)
            st.rerun()

    with col2:
        if st.button("Limpar LM", use_container_width=True):
            st.session_state.material_list_items = []
            st.rerun()

    with col3:
        lm_data = {
            "lm_number": lm_number,
            "issue_date": issue_date,
            "deadline": deadline,
            "issued_by": issued_by,
            "client_project": client_project,
            "comments": comments,
        }
        excel_file = build_excel(lm_data, items_df)
        clean_number = lm_number.strip().replace("/", "-").replace("\\", "-") or "sem_numero"

        st.download_button(
            "Baixar LM em Excel",
            data=excel_file,
            file_name=f"LM_{clean_number}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
